import os
import openpyxl
import re
from datetime import datetime
from pathlib import Path
from telegram import Update, ReplyKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes, ConversationHandler
import logging
import tempfile
from collections import defaultdict

# Настройка логирования
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Состояния для ConversationHandler
SELECT_TYPE, INPUT_VALUE = range(2)

# Глобальная переменная для хранения данных
user_data_dict = {}

def get_current_month():
    """Возвращает текущий месяц на русском"""
    month_names = {
        '01': 'январь', '02': 'февраль', '03': 'март', '04': 'апрель',
        '05': 'май', '06': 'июнь', '07': 'июль', '08': 'август',
        '09': 'сентябрь', '10': 'октябрь', '11': 'ноябрь', '12': 'декабрь'
    }
    current_month = datetime.now().month
    return month_names.get(str(current_month).zfill(2))

def find_table_structure(ws):
    """Находит структуру таблицы по ключевым заголовкам"""
    headers_positions = {}
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                cell_value = str(cell.value).strip()
                
                if "Товары (работы, услуги)" in cell_value:
                    headers_positions['description'] = (cell.row, cell.column)
                elif "Сумма" in cell_value and cell_value != "Сумма с НДС":
                    headers_positions['amount'] = (cell.row, cell.column)
                elif "№" == cell_value and cell.column < 10:
                    headers_positions['number'] = (cell.row, cell.column)
                elif "Кол-во" in cell_value:
                    headers_positions['quantity'] = (cell.row, cell.column)
                elif "Ед." in cell_value:
                    headers_positions['unit'] = (cell.row, cell.column)
                elif "Цена" in cell_value:
                    headers_positions['price'] = (cell.row, cell.column)
    
    return headers_positions

def extract_data_from_description(description):
    """Извлекает дату, маршрут, гос. номер и фамилию водителя из описания"""
    description_str = str(description)
    
    # Маршрут (все до первой запятой)
    route = description_str.split(',')[0].strip()
    
    # Дата из текста (формат "от 06.09.25")
    date_match = re.search(r'от\s+(\d{2}\.\d{2}\.\d{2})', description_str)
    date_str = date_match.group(1) if date_match else "Дата не найдена"
    
    # Гос. номер - ищем 3 цифры подряд
    plate_match = re.search(r'(\d{3})', description_str)
    car_plate = plate_match.group(1) if plate_match else "Неизвестно"
    
    # Фамилия водителя
    driver_match = re.search(r',\s*([А-Я][а-я]+)\s+[А-Я]\.[А-Я]\.', description_str)
    if driver_match:
        driver_name = driver_match.group(1)
    else:
        alt_driver_match = re.search(r',\s*([А-Я][а-я]+)', description_str)
        driver_name = alt_driver_match.group(1) if alt_driver_match else "Фамилия не найдена"
    
    return route, date_str, car_plate, driver_name

def parse_invoice_file(file_path):
    """Парсит один файл счета и возвращает данные"""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        headers = find_table_structure(ws)
        
        if not headers.get('description') or not headers.get('amount'):
            return []
        
        header_row = max(h[0] for h in headers.values())
        description_col = headers['description'][1]
        amount_col = headers['amount'][1]
        
        parsed_data = []
        row_num = header_row + 1
        processed_count = 0
        current_empty_rows = 0
        max_empty_rows = 5
        
        while current_empty_rows < max_empty_rows:
            description_cell = ws.cell(row=row_num, column=description_col)
            description = description_cell.value
            
            if not description:
                current_empty_rows += 1
                row_num += 1
                continue
                
            current_empty_rows = 0
            description_str = str(description)
            
            if any(word in description_str.lower() for word in ['итого', 'всего', 'итог', 'сумма']):
                row_num += 1
                continue
            
            amount_cell = ws.cell(row=row_num, column=amount_col)
            amount = amount_cell.value
            
            if amount is not None:
                try:
                    if isinstance(amount, str) and any(char.isalpha() for char in amount.replace(' ', '').replace(',', '.')):
                        row_num += 1
                        continue
                    
                    amount_str = str(amount).replace(' ', '').replace(',', '.')
                    amount_value = float(amount_str)
                    
                    route, date_str, car_plate, driver_name = extract_data_from_description(description_str)
                    
                    if car_plate != "Неизвестно" and amount_value > 0:
                        parsed_data.append({
                            'Дата': date_str,
                            'Маршрут': route,
                            'Стоимость': amount_value,
                            'Гос_номер': car_plate,
                            'Водитель': driver_name,
                            'Источник': file_path.name
                        })
                        processed_count += 1
                    
                except (ValueError, TypeError):
                    pass
            
            row_num += 1
            
            if row_num > header_row + 1000:
                break
        
        return parsed_data
        
    except Exception as e:
        logger.error(f"Ошибка при обработке файла: {e}")
        return []

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /start"""
    user = update.message.from_user
    welcome_text = (
        f"Привет, {user.first_name}! 🚛\n\n"
        "Я бот для анализа транспортных отчетов.\n\n"
        "📊 Что я умею:\n"
        "• Анализировать Excel-файлы с поездками\n"
        "• Показывать общую статистику\n"
        "• Искать данные по гос. номеру или водителю\n\n"
        "📎 Просто отправь мне файл Excel с отчетом!\n\n"
        "Команды:\n"
        "/stats - общая статистика\n"
        "/search - поиск по номеру или водителю\n"
        "/clear - очистить данные\n"
        "/help - справка"
    )
    await update.message.reply_text(welcome_text)

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /help"""
    help_text = (
        "📋 Как пользоваться ботом:\n\n"
        "1. 📎 Отправь Excel-файл с отчетом о поездках\n"
        "2. 📊 Используй /stats для просмотра статистики\n"
        "3. 🔍 Используй /search для поиска по:\n"
        "   • Гос. номеру (например, 123)\n"
        "   • Фамилии водителя\n"
        "4. 🗑️ /clear - очистить все данные\n\n"
        "Формат файлов: стандартные Excel-файлы с маршрутами, суммами, гос. номерами и водителями."
    )
    await update.message.reply_text(help_text)

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик загрузки документов"""
    user_id = update.message.from_user.id
    
    try:
        # Создаем папку для пользователя, если её нет
        if user_id not in user_data_dict:
            user_data_dict[user_id] = []
        
        # Скачиваем файл
        document = update.message.document
        file = await context.bot.get_file(document.file_id)
        
        # Создаем временный файл
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            await file.download_to_drive(temp_file.name)
            
            # Парсим файл
            parsed_data = parse_invoice_file(Path(temp_file.name))
            
            # Добавляем данные пользователю
            user_data_dict[user_id].extend(parsed_data)
            
            # Удаляем временный файл
            os.unlink(temp_file.name)
        
        # Формируем ответ
        total_files = len(set(item['Источник'] for item in user_data_dict[user_id]))
        total_trips = len(parsed_data)
        total_amount = sum(item['Стоимость'] for item in parsed_data)
        unique_cars = len(set(item['Гос_номер'] for item in parsed_data))
        unique_drivers = len(set(item['Водитель'] for item in parsed_data))
        
        response = (
            f"✅ Файл успешно обработан!\n\n"
            f"📊 Статистика файла:\n"
            f"• Поездок: {total_trips}\n"
            f"• Сумма: {total_amount:,.0f} руб.\n"
            f"• Уникальных машин: {unique_cars}\n"
            f"• Уникальных водителей: {unique_drivers}\n\n"
            f"📁 Всего файлов: {total_files}\n"
            f"📈 Всего поездок: {len(user_data_dict[user_id])}\n\n"
            f"Используй /stats для полной статистики или /search для поиска."
        )
        
        await update.message.reply_text(response)
        
    except Exception as e:
        logger.error(f"Ошибка обработки файла: {e}")
        await update.message.reply_text("❌ Ошибка при обработке файла. Убедитесь, что это корректный Excel-файл с отчетами.")

def calculate_statistics(data):
    """Рассчитывает статистику без pandas"""
    if not data:
        return {}
    
    # Общая статистика
    total_trips = len(data)
    total_amount = sum(item['Стоимость'] for item in data)
    
    # Статистика по автомобилям
    car_stats = defaultdict(lambda: {'sum': 0, 'count': 0, 'drivers': set()})
    for item in data:
        if item['Гос_номер'] != "Неизвестно":
            car_stats[item['Гос_номер']]['sum'] += item['Стоимость']
            car_stats[item['Гос_номер']]['count'] += 1
            car_stats[item['Гос_номер']]['drivers'].add(item['Водитель'])
    
    # Статистика по водителям
    driver_stats = defaultdict(lambda: {'sum': 0, 'count': 0, 'cars': set()})
    for item in data:
        if item['Водитель'] != "Фамилия не найдена":
            driver_stats[item['Водитель']]['sum'] += item['Стоимость']
            driver_stats[item['Водитель']]['count'] += 1
            driver_stats[item['Водитель']]['cars'].add(item['Гос_номер'])
    
    # Уникальные значения
    unique_cars = len([car for car in car_stats.keys() if car != "Неизвестно"])
    unique_drivers = len([driver for driver in driver_stats.keys() if driver != "Фамилия не найдена"])
    unique_files = len(set(item['Источник'] for item in data))
    
    return {
        'total_trips': total_trips,
        'total_amount': total_amount,
        'unique_cars': unique_cars,
        'unique_drivers': unique_drivers,
        'unique_files': unique_files,
        'car_stats': dict(car_stats),
        'driver_stats': dict(driver_stats)
    }

async def show_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показать общую статистику"""
    user_id = update.message.from_user.id
    
    if user_id not in user_data_dict or not user_data_dict[user_id]:
        await update.message.reply_text("📭 Нет данных для анализа. Сначала отправьте файлы с отчетами.")
        return
    
    stats = calculate_statistics(user_data_dict[user_id])
    
    # Формируем текст статистики по автомобилям
    car_stats_text = "🚗 Статистика по автомобилям:\n"
    for car_plate, car_data in stats['car_stats'].items():
        if car_plate != "Неизвестно":
            amount = car_data['sum']
            count = car_data['count']
            drivers = len(car_data['drivers'])
            car_stats_text += f"• {car_plate}: {count} поездок, {amount:,.0f} руб., {drivers} водит.\n"
    
    # Формируем текст статистики по водителям
    driver_stats_text = "\n👤 Статистика по водителям:\n"
    for driver, driver_data in stats['driver_stats'].items():
        if driver != "Фамилия не найдена":
            amount = driver_data['sum']
            count = driver_data['count']
            cars = len(driver_data['cars'])
            driver_stats_text += f"• {driver}: {count} поездок, {amount:,.0f} руб., {cars} машин\n"
    
    response = (
        f"📊 ОБЩАЯ СТАТИСТИКА\n\n"
        f"📈 Основные показатели:\n"
        f"• Всего поездок: {stats['total_trips']}\n"
        f"• Общая сумма: {stats['total_amount']:,.0f} руб.\n"
        f"• Автомобилей: {stats['unique_cars']}\n"
        f"• Водителей: {stats['unique_drivers']}\n"
        f"• Файлов: {stats['unique_files']}\n\n"
        f"{car_stats_text}"
        f"{driver_stats_text}"
    )
    
    await update.message.reply_text(response)

async def search_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Начало поиска"""
    user_id = update.message.from_user.id
    
    if user_id not in user_data_dict or not user_data_dict[user_id]:
        await update.message.reply_text("📭 Нет данных для поиска. Сначала отправьте файлы с отчетами.")
        return ConversationHandler.END
    
    keyboard = [['🚗 По гос. номеру', '👤 По водителю']]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True)
    
    await update.message.reply_text(
        "🔍 Выберите тип поиска:",
        reply_markup=reply_markup
    )
    
    return SELECT_TYPE

async def select_search_type(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка выбора типа поиска"""
    search_type = update.message.text
    
    if search_type == '🚗 По гос. номеру':
        context.user_data['search_type'] = 'car'
        await update.message.reply_text(
            "Введите гос. номер (только цифры, например: 123):",
            reply_markup=None
        )
    elif search_type == '👤 По водителю':
        context.user_data['search_type'] = 'driver'
        await update.message.reply_text(
            "Введите фамилию водителя:",
            reply_markup=None
        )
    else:
        await update.message.reply_text("Пожалуйста, выберите тип поиска из предложенных вариантов.")
        return SELECT_TYPE
    
    return INPUT_VALUE

async def perform_search(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Выполнение поиска"""
    user_id = update.message.from_user.id
    search_value = update.message.text
    search_type = context.user_data.get('search_type')
    
    data = user_data_dict[user_id]
    
    if search_type == 'car':
        results = [item for item in data if item['Гос_номер'] == search_value]
        search_title = f"🚗 Результаты поиска по гос. номеру: {search_value}"
    else:
        results = [item for item in data if search_value.lower() in item['Водитель'].lower()]
        search_title = f"👤 Результаты поиска по водителю: {search_value}"
    
    if not results:
        await update.message.reply_text(f"❌ По вашему запросу ничего не найдено.")
        return ConversationHandler.END
    
    total_trips = len(results)
    total_amount = sum(item['Стоимость'] for item in results)
    avg_amount = total_amount / total_trips if total_trips > 0 else 0
    
    # Детализация поездок
    details_text = "\n📋 Последние поездки:\n"
    for item in results[:10]:  # Показываем первые 10
        details_text += f"• {item['Дата']}: {item['Маршрут'][:30]}... - {item['Стоимость']:,.0f} руб.\n"
    
    if len(results) > 10:
        details_text += f"... и еще {len(results) - 10} поездок\n"
    
    response = (
        f"{search_title}\n\n"
        f"📊 Статистика:\n"
        f"• Количество поездок: {total_trips}\n"
        f"• Общая сумма: {total_amount:,.0f} руб.\n"
        f"• Средняя стоимость: {avg_amount:,.0f} руб.\n"
        f"{details_text}"
    )
    
    await update.message.reply_text(response)
    return ConversationHandler.END

async def cancel_search(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Отмена поиска"""
    await update.message.reply_text("Поиск отменен.")
    return ConversationHandler.END

async def clear_data(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Очистка всех данных пользователя"""
    user_id = update.message.from_user.id
    
    if user_id in user_data_dict:
        user_data_dict[user_id] = []
        await update.message.reply_text("✅ Все данные очищены.")
    else:
        await update.message.reply_text("📭 Нет данных для очистки.")

async def error_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик ошибок"""
    logger.error(f"Ошибка: {context.error}")
    if update and update.message:
        await update.message.reply_text("❌ Произошла ошибка. Попробуйте еще раз.")

def main():
    """Запуск бота"""
    # Получаем токен из переменных окружения Render
    TOKEN = os.getenv('BOT_TOKEN')
    
    if not TOKEN:
        logger.error("Токен бота не найден в переменных окружения!")
        return
    
    # Создаем приложение
    application = Application.builder().token(TOKEN).build()
    
    # Обработчики команд
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("help", help_command))
    application.add_handler(CommandHandler("stats", show_stats))
    application.add_handler(CommandHandler("clear", clear_data))
    
    # Обработчик документов
    application.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    
    # ConversationHandler для поиска
    search_conv = ConversationHandler(
        entry_points=[CommandHandler("search", search_start)],
        states={
            SELECT_TYPE: [MessageHandler(filters.TEXT & ~filters.COMMAND, select_search_type)],
            INPUT_VALUE: [MessageHandler(filters.TEXT & ~filters.COMMAND, perform_search)],
        },
        fallbacks=[CommandHandler("cancel", cancel_search)]
    )
    application.add_handler(search_conv)
    
    # Обработчик ошибок
    application.add_error_handler(error_handler)
    
    # Запуск бота
    print("Бот запущен...")
    application.run_polling()

if __name__ == "__main__":
    main()
