import pandas as pd
import openpyxl
import re
import io # Добавляем io для работы с файлами в памяти

def find_table_structure(ws):
    """Находит структуру таблицы по ключевым заголовкам"""
    headers_positions = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                cell_value = str(cell.value).strip()
                if "Товары (работы, услуги)" in cell_value:
                    headers_positions['description'] = (cell.row, cell.column)
                elif "Сумма" in cell_value and "Сумма с НДС" not in cell_value:
                    headers_positions['amount'] = (cell.row, cell.column)
    return headers_positions

def extract_data_from_description(description):
    """Извлекает дату, маршрут, гос. номер и фамилию водителя из описания"""
    description_str = str(description)
    
    route = description_str.split(',')[0].strip()
    
    date_match = re.search(r'от\s+(\d{2}\.\d{2}\.\d{2})', description_str)
    date_str = date_match.group(1) if date_match else "Дата не найдена"
    
    plate_match = re.search(r'(\d{3})', description_str)
    car_plate = plate_match.group(1) if plate_match else "Неизвестно"
    
    driver_match = re.search(r',\s*([А-ЯЁ][а-яё]+)\s+[А-ЯЁ]\.[А-ЯЁ]\.', description_str)
    if driver_match:
        driver_name = driver_match.group(1)
    else:
        alt_driver_match = re.search(r',\s*([А-ЯЁ][а-яё]+)', description_str)
        driver_name = alt_driver_match.group(1) if alt_driver_match else "Фамилия не найдена"
    
    return route, date_str, car_plate, driver_name

def process_excel_file(file_content: bytes, file_name: str):
    """
    Парсит один Excel-файл из байтового потока и возвращает DataFrame.
    """
    try:
        # Используем io.BytesIO для чтения файла из памяти
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        ws = wb.active
        
        headers = find_table_structure(ws)
        
        if not headers.get('description') or not headers.get('amount'):
            print(f"⚠️ В файле {file_name} не найдена структура таблицы.")
            return None
        
        header_row = max(h[0] for h in headers.values())
        description_col = headers['description'][1]
        amount_col = headers['amount'][1]
        
        parsed_data = []
        row_num = header_row + 1
        
        # Проходимся по строкам до определенного предела или пока есть данные
        for row_num in range(header_row + 1, ws.max_row + 1):
            description = ws.cell(row=row_num, column=description_col).value
            amount = ws.cell(row=row_num, column=amount_col).value

            if not description or not amount:
                continue

            description_str = str(description)
            if any(word in description_str.lower() for word in ['итого', 'всего', 'сумма']):
                continue

            try:
                amount_str = str(amount).replace(' ', '').replace(',', '.')
                amount_value = float(amount_str)

                route, date_str, car_plate, driver_name = extract_data_from_description(description_str)
                
                if car_plate != "Неизвестно" and amount_value > 0:
                    parsed_data.append({
                        'Дата': date_str,
                        'Маршрут': route,
                        'Стоимость': amount_value,
                        'Гос_номер': car_plate,
                        'Водитель': driver_name,
                        'Источник': file_name
                    })
            except (ValueError, TypeError):
                continue
        
        if not parsed_data:
            return None

        return pd.DataFrame(parsed_data)

    except Exception as e:
        print(f"❌ Ошибка при обработке файла {file_name}: {e}")
        return None
